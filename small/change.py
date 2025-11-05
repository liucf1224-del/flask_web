from openpyxl import load_workbook
import os


def replace_chinese_in_excel(input_file, output_file, replacements):
    """
    在Excel文件中批量替换中文文本
    参数:
    input_file: 输入的Excel文件路径
    output_file: 输出的Excel文件路径
    replacements: 替换字典 {旧文本: 新文本}
    """
    # 加载工作簿
    wb = load_workbook(input_file)

    # 遍历所有工作表
    for sheet in wb:
        # 遍历工作表中的所有单元格
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # 对每个替换项执行替换操作
                    for old_text, new_text in replacements.items():
                        # 使用中文安全替换
                        if old_text in cell.value:
                            cell.value = cell.value.replace(old_text, new_text)

    # 保存修改后的工作簿
    wb.save(output_file)
    print(f"中文替换完成! 新文件保存至: {os.path.abspath(output_file)}")


if __name__ == "__main__":
    # 配置替换参数
    replacements = {
        # "个旧市恺撒歌厅": "南明区尚城悦色娱乐中心",
        # "个旧市恺撒歌厅": "襄阳百金瀚宫娱乐有限公司",
        # "个旧市恺撒歌厅": "锦煌夜都",
        # "个旧市恺撒歌厅": "师宗县黑桃壹娱乐俱乐部",
        # "个旧市恺撒歌厅": "云南恋在俪江娱乐",
        # "个旧市恺撒歌厅": "永胜县三川镇鑫丽派对量贩",
        # "个旧市恺撒歌厅": "贵州元一娱乐会所",
        # "个旧市恺撒歌厅": "昭通市爱尚娱乐有限公司",
        # "个旧市恺撒歌厅": "印江宇宙空间",
        # "个旧市恺撒歌厅": "楚雄开发区天逸华庭娱乐城",
        # "个旧市恺撒歌厅": "鹰潭福纳娱乐有限公司",
        # "个旧市恺撒歌厅": "陆良紫景文化",
        # "个旧市恺撒歌厅": "绿宝银河量贩",
        # "个旧市恺撒歌厅": "蒙自永利银河汇娱乐中心",
        # "个旧市恺撒歌厅": "富民县彼岸花香歌舞厅",
        # "个旧市恺撒歌厅": "云南光辉岁月文化传播有限公司",
        # "个旧市恺撒歌厅": "云龙县诺邓镇云上派对歌厅",
        # "个旧市恺撒歌厅": "维西潮歌娱乐城",
        # "个旧市恺撒歌厅": "瑞丽市天易顺娱乐",
        # "个旧市恺撒歌厅": "蒙自市缘顶文化",
        # "个旧市恺撒歌厅": "保山市隆阳区东悦娱乐",
        # "个旧市恺撒歌厅": "上饶市万歌汇娱乐有限公司",
        # "个旧市恺撒歌厅": "禄劝屏山镇好声音歌城",
        # "个旧市恺撒歌厅": "沧源县鼎渝商贸有限公司",
        "个旧市恺撒歌厅": "武定麦颂娱乐中心",
        # 可以添加更多中文替换项:
        # "旧文本2": "新文本2",
    }

    # 文件路径配置
    input_excel = r"D:\模板\模板批量导入 -凯撒 - 副本 (6).xlsx"  # 替换为你的Excel文件路径
    first_new_name = list(replacements.values())[0]
    output_excel = fr"D:\模板\change\模板批量导入-{first_new_name}.xlsx"  # 新文件名


    # 执行替换
    replace_chinese_in_excel(input_excel, output_excel, replacements)
    print(f"已将所有 '{list(replacements.keys())[0]}' 替换为 '{list(replacements.values())[0]}'")